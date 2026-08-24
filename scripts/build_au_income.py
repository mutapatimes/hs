"""Build the Australia postcode income table from ATO open taxation statistics.

The Australian Taxation Office publishes annual Taxation Statistics on data.gov.au,
including Individuals Table 6B: taxable income by postcode for the whole country. This
script downloads the workbook (or takes an already-downloaded copy), computes the AVERAGE
TAXABLE INCOME per postcode, assigns wealth tiers, and writes
reference_data/postcodes/au_income_values.csv, which the au_income signal reads.

The table is national and complete, so the output is fully regenerated each run (no curated
merge, unlike the property builds). Postcodes with few earners and PO-box/LVR ranges are
excluded: a GPO-box postcode is a mail population, not a residential one.

Needs openpyxl (in the project venv). Ingest offline, never a live lookup on a customer.

Usage
-----
    # Download the latest published year and rebuild:
    python scripts/build_au_income.py

    # Already-downloaded workbook(s):
    python scripts/build_au_income.py --files ~/Downloads/ts24individual06*.xlsx
"""
from __future__ import annotations

import argparse
import csv
import sys
import tempfile
import urllib.error
import urllib.request
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
from config import AU_INCOME_VALUES_FILE, AU_PROPERTY_VALUES_FILE  # noqa: E402

# Tier bands in AUD (postcode average taxable income). Calibrated on the 2023-24 national
# distribution: median postcode averages ~$70k; ultra catches ~11 postcodes (Portsea $330k,
# Toorak $282k, Bellevue Hill $253k...), prime ~34, high ~100.
ULTRA = 200_000
PRIME = 150_000
HIGH = 120_000

MIN_EARNERS = 500   # thin postcodes give noisy averages

# Large-volume-receiver / PO-box postcode ranges: high figures there are mailrooms, not homes.
_PO_BOX_RANGES = [(1000, 1999), (5800, 5999), (6800, 6999), (7800, 7999),
                  (8000, 8999), (9000, 9999), (200, 299)]

# ATO Taxation Statistics 2023-24, Individuals Table 6 (state/SA4/postcode), data.gov.au.
_URL = ("https://data.gov.au/data/dataset/faea4485-f407-457d-97f8-3f0822ccd654/resource/"
        "9d8577b7-a096-4758-9b3b-0649f3b83de7/download/"
        "ts24individual06taxablestatusstatesa4postcode.xlsx")

_SHEET = "Table 6B"
_STATE_IX, _SA4_IX, _PC_IX, _EARNERS_IX, _TOTAL_IX = 0, 1, 2, 4, 5


def _tier(avg: float) -> str | None:
    if avg >= ULTRA:
        return "ultra"
    if avg >= PRIME:
        return "prime"
    if avg >= HIGH:
        return "high"
    return None


def _is_po_box(pc: str) -> bool:
    n = int(pc)
    return any(lo <= n <= hi for lo, hi in _PO_BOX_RANGES)


def _property_names() -> dict[str, str]:
    """Locality names from the property seed, so reasons read 'Toorak' not 'Melbourne - Inner'."""
    names: dict[str, str] = {}
    path = Path(AU_PROPERTY_VALUES_FILE)
    if not path.exists():
        return names
    with path.open(newline="", encoding="utf-8") as fh:
        for row in csv.reader(fh):
            if row and row[0].strip().isdigit() and len(row) >= 2 and row[1].strip():
                names[row[0].strip()] = row[1].strip()
    return names


def _ingest(path: Path, out: dict[str, dict]) -> int:
    wb = openpyxl.load_workbook(path, read_only=True)
    if _SHEET not in wb.sheetnames:
        print(f"{path}: no '{_SHEET}' sheet", file=sys.stderr)
        return 0
    used = 0
    for i, row in enumerate(wb[_SHEET].iter_rows(values_only=True)):
        if i < 2 or row is None or len(row) <= _TOTAL_IX:
            continue
        pc = str(row[_PC_IX] or "").strip()
        if not (len(pc) == 4 and pc.isdigit()) or _is_po_box(pc):
            continue
        try:
            earners = float(row[_EARNERS_IX] or 0)
            total = float(row[_TOTAL_IX] or 0)
        except (TypeError, ValueError):
            continue
        if earners < MIN_EARNERS or total <= 0:
            continue
        out[pc] = {"avg": total / earners, "sa4": str(row[_SA4_IX] or "").strip()}
        used += 1
    return used


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    ap.add_argument("--files", nargs="+", type=Path, default=None,
                    help="Already-downloaded ATO Individuals Table 6 workbook(s)")
    ap.add_argument("--out", type=Path, default=Path(AU_INCOME_VALUES_FILE))
    args = ap.parse_args()

    stats: dict[str, dict] = {}
    with tempfile.TemporaryDirectory() as td:
        files = args.files or []
        if not files:
            dest = Path(td) / "ato_individuals_t6.xlsx"
            print(f"Downloading {_URL} …")
            try:
                urllib.request.urlretrieve(_URL, dest)  # noqa: S310 — fixed https host
            except urllib.error.URLError as exc:
                print(f"Could not download: {exc}", file=sys.stderr)
                return 1
            files = [dest]
        total = 0
        for f in files:
            n = _ingest(Path(f), stats)
            print(f"{f}: {n:,} postcodes read")
            total += n
    if not total:
        print("No usable rows found.", file=sys.stderr)
        return 1

    names = _property_names()
    rows = []
    for pc, s in stats.items():
        tier = _tier(s["avg"])
        if tier:
            rows.append([pc, names.get(pc, s["sa4"]), str(int(s["avg"])), tier])
    rows.sort(key=lambda r: r[0])

    args.out.parent.mkdir(parents=True, exist_ok=True)
    with args.out.open("w", newline="", encoding="utf-8") as fh:
        w = csv.writer(fh)
        w.writerow(["postcode", "area", "avg_taxable_income", "tier"])
        fh.write("# Generated by scripts/build_au_income.py from ATO Taxation Statistics (data.gov.au).\n")
        fh.write(f"# Tier bands (AUD average taxable income): ultra>={ULTRA:,} prime>={PRIME:,} "
                 f"high>={HIGH:,}; min {MIN_EARNERS} earners; PO-box ranges excluded.\n")
        w.writerows(rows)
    print(f"Wrote {len(rows):,} rows to {args.out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
